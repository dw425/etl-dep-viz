"""Excel workbook export — multi-sheet openpyxl workbook (Item 82)."""

from __future__ import annotations

import io
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def generate_excel_workbook(tier_data: dict, vector_results: dict | None = None) -> bytes:
    """Generate a multi-sheet Excel workbook from tier data.

    Sheets:
      1. Sessions — all session details
      2. Tables — table catalog with types and conflict counts
      3. Connections — dependency edges
      4. Statistics — summary stats
      5. Complexity — V11 scores (if available)
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0'),
    )

    def write_header(ws, headers: list[str]):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # ── Sheet 1: Sessions ──
    ws_sessions = wb.active
    ws_sessions.title = "Sessions"
    session_headers = ["ID", "Name", "Full Name", "Tier", "Step", "Transforms", "Ext Reads", "Lookups", "Critical"]
    write_header(ws_sessions, session_headers)
    for i, s in enumerate(tier_data.get('sessions', []), 2):
        ws_sessions.cell(row=i, column=1, value=s.get('id', ''))
        ws_sessions.cell(row=i, column=2, value=s.get('name', ''))
        ws_sessions.cell(row=i, column=3, value=s.get('full', ''))
        ws_sessions.cell(row=i, column=4, value=s.get('tier', 0))
        ws_sessions.cell(row=i, column=5, value=s.get('step', 0))
        ws_sessions.cell(row=i, column=6, value=s.get('transforms', 0))
        ws_sessions.cell(row=i, column=7, value=s.get('extReads', 0))
        ws_sessions.cell(row=i, column=8, value=s.get('lookupCount', 0))
        ws_sessions.cell(row=i, column=9, value='Yes' if s.get('critical') else 'No')

    # Auto-fit columns
    for col in ws_sessions.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_sessions.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ── Sheet 2: Tables ──
    ws_tables = wb.create_sheet("Tables")
    table_headers = ["ID", "Name", "Type", "Tier", "Conflict Writers", "Readers", "Lookup Users"]
    write_header(ws_tables, table_headers)
    for i, t in enumerate(tier_data.get('tables', []), 2):
        ws_tables.cell(row=i, column=1, value=t.get('id', ''))
        ws_tables.cell(row=i, column=2, value=t.get('name', ''))
        ws_tables.cell(row=i, column=3, value=t.get('type', ''))
        ws_tables.cell(row=i, column=4, value=t.get('tier', 0))
        ws_tables.cell(row=i, column=5, value=t.get('conflictWriters', 0))
        ws_tables.cell(row=i, column=6, value=t.get('readers', 0))
        ws_tables.cell(row=i, column=7, value=t.get('lookupUsers', 0))

    for col in ws_tables.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_tables.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ── Sheet 3: Connections ──
    ws_conns = wb.create_sheet("Connections")
    conn_headers = ["From", "To", "Type"]
    write_header(ws_conns, conn_headers)
    for i, c in enumerate(tier_data.get('connections', []), 2):
        ws_conns.cell(row=i, column=1, value=c.get('from', ''))
        ws_conns.cell(row=i, column=2, value=c.get('to', ''))
        ws_conns.cell(row=i, column=3, value=c.get('type', ''))

    # ── Sheet 4: Statistics ──
    ws_stats = wb.create_sheet("Statistics")
    write_header(ws_stats, ["Metric", "Value"])
    stats = tier_data.get('stats', {})
    stat_rows = [
        ("Session Count", stats.get('session_count', 0)),
        ("Write Conflicts", stats.get('write_conflicts', 0)),
        ("Dependency Chains", stats.get('dep_chains', 0)),
        ("Staleness Risks", stats.get('staleness_risks', 0)),
        ("Source Tables", stats.get('source_tables', 0)),
        ("Max Tier Depth", stats.get('max_tier', 0)),
    ]
    for i, (metric, value) in enumerate(stat_rows, 2):
        ws_stats.cell(row=i, column=1, value=metric)
        ws_stats.cell(row=i, column=2, value=value)
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 15

    # ── Sheet 5: Complexity (if vector results available) ──
    if vector_results and 'v11_complexity' in vector_results:
        ws_complex = wb.create_sheet("Complexity")
        complex_headers = ["Session ID", "Overall Score", "Bucket", "Transform Score", "IO Score", "Dependency Score"]
        write_header(ws_complex, complex_headers)
        scores = vector_results['v11_complexity'].get('scores', [])
        for i, s in enumerate(scores, 2):
            ws_complex.cell(row=i, column=1, value=s.get('session_id', ''))
            ws_complex.cell(row=i, column=2, value=s.get('overall_score', 0))
            ws_complex.cell(row=i, column=3, value=s.get('bucket', ''))
            ws_complex.cell(row=i, column=4, value=s.get('transform_score', 0))
            ws_complex.cell(row=i, column=5, value=s.get('io_score', 0))
            ws_complex.cell(row=i, column=6, value=s.get('dependency_score', 0))

        for col in ws_complex.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_complex.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
